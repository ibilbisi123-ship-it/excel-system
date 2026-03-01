import argparse
import os
import re
import sqlite3
from openpyxl import load_workbook


NUMERIC_RE = re.compile(r"^\d+(?:\.\d+)?$")


def normalize_mhr(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if NUMERIC_RE.fullmatch(text):
            return float(text)
    return None


def has_alnum(text):
    return bool(re.search(r"[A-Za-z0-9]", text))


def find_header(ws, max_rows=30):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        desc_col = None
        mhr_col = None
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str):
                continue
            key = val.strip().lower()
            if key == "description":
                desc_col = c
            elif key == "mhr":
                mhr_col = c
        if desc_col and mhr_col:
            return r, desc_col, mhr_col
    return None, None, None


def import_xlsx_to_db(xlsx_path, db_path, table):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header_row, desc_col, mhr_col = find_header(ws)
    if not header_row:
        raise ValueError("Could not find header row with Description and Mhr columns.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        if desc is None:
            continue
        desc_text = str(desc).strip()
        if not desc_text or not has_alnum(desc_text):
            continue
        mhr_val = normalize_mhr(ws.cell(row=r, column=mhr_col).value)
        if mhr_val is None:
            continue
        rows.append((desc_text, mhr_val, mhr_val))

    if not rows:
        raise ValueError("No valid Description/Mhr rows found to import.")

    conn = sqlite3.connect(db_path)
    try:
        cur = conn.cursor()
        cur.execute(f'DROP TABLE IF EXISTS "{table}"')
        cur.execute(
            f'CREATE TABLE "{table}" ('
            'description TEXT, '
            '"Cable laying Mhr" REAL, '
            '"Cable termination Mhr" REAL'
            ")"
        )
        cur.executemany(
            f'INSERT INTO "{table}" (description, "Cable laying Mhr", "Cable termination Mhr") VALUES (?, ?, ?)',
            rows,
        )
        conn.commit()
    finally:
        conn.close()

    return len(rows)


def main():
    parser = argparse.ArgumentParser(description="Import equipment Mhr Excel into SQLite DB.")
    parser.add_argument("--xlsx", default="equipment mhr (1).xlsx", help="Path to source .xlsx file.")
    parser.add_argument("--db", default="my_database.db", help="Path to SQLite database to create.")
    parser.add_argument("--table", default="my_table", help="Target table name.")
    args = parser.parse_args()

    xlsx_path = os.path.abspath(args.xlsx)
    db_path = os.path.abspath(args.db)
    count = import_xlsx_to_db(xlsx_path, db_path, args.table)
    print(f"Imported {count} rows into {db_path} ({args.table}).")


if __name__ == "__main__":
    main()
