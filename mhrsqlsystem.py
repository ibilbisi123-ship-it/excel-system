# estimator_tk_preserve_layout_sheetpicker_unified_only.py
# Modernized Tkinter app with improved UI/UX - FIXED PROGRESS BAR
# This version ONLY merges the three step frames into a single background frame.
# All other code is kept identical to the user's original.

import re
import os
import threading
import sqlite3
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from tkinter.font import Font

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ===================== CONFIG: set your credentials =====================
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")  # Set via environment variable
GPT_MODEL = "gpt-4.1-mini"
DB_FILENAME = "my_database.db"
DB_TABLE = "my_table"
# Base directory to locate the local database (supports execution from IDE or packaged app)
BASE_DIR = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.getcwd()
DB_PATH = os.path.join(BASE_DIR, DB_FILENAME)
MAX_DB_CANDIDATES = 18
MAX_GPT_CONTEXT = 12
MIN_CANDIDATE_SCORE = 0.2
MIN_RESULTS_BEFORE_THRESHOLD = 5
# ========================================================================

client = OpenAI(api_key=OPENAI_API_KEY)

# Modern color scheme
COLORS = {
    "primary": "#2563eb",
    "secondary": "#64748b",
    "success": "#22c55e",
    "warning": "#f59e0b",
    "error": "#ef4444",
    "dark_bg": "#0f172a",
    "light_bg": "#1e293b",
    "card_bg": "#1e293b",
    "text_primary": "#f1f5f9",
    "text_secondary": "#94a3b8",
    "border": "#334155"
}


TOKEN_PATTERN = re.compile(r"[a-z0-9]+(?:[/-][a-z0-9]+)*")
NUMBER_PATTERN = re.compile(r"\d+(?:\.\d+)?")
STOPWORDS = {"and", "or", "the", "to", "with", "without", "for", "in", "of", "a", "an", "per"}


def normalize_description(text: str) -> str:
    if text is None:
        return ""
    return " ".join(str(text).lower().split())


def extract_tokens(text: str) -> list:
    if not text:
        return []
    raw = str(text).lower()
    base_tokens = []
    for match in TOKEN_PATTERN.finditer(raw):
        token = match.group()
        if token in STOPWORDS:
            continue
        base_tokens.append(token)
        if "-" in token or "/" in token:
            for part in re.split(r"[-/]", token):
                if part and part not in STOPWORDS:
                    base_tokens.append(part)
    enriched_tokens = []
    for token in base_tokens:
        if token in STOPWORDS:
            continue
        enriched_tokens.append(token)
        alnum = re.match(r"^(\d+)([a-z]+)$", token)
        if alnum:
            enriched_tokens.append(alnum.group(1))
            suffix = alnum.group(2)
            if suffix and suffix not in STOPWORDS:
                enriched_tokens.append(suffix)
        else:
            rev = re.match(r"^([a-z]+)(\d+)$", token)
            if rev:
                prefix = rev.group(1)
                suffix = rev.group(2)
                if prefix and prefix not in STOPWORDS:
                    enriched_tokens.append(prefix)
                enriched_tokens.append(suffix)
    for idx in range(len(base_tokens) - 1):
        first = base_tokens[idx]
        second = base_tokens[idx + 1]
        if first.isdigit() and second.isalpha() and second not in STOPWORDS:
            enriched_tokens.append(f"{first}{second}")
        if second.isdigit() and first.isalpha() and first not in STOPWORDS:
            enriched_tokens.append(f"{first}{second}")
    unique = []
    for token in enriched_tokens:
        if token and token not in unique and token not in STOPWORDS:
            unique.append(token)
    return unique


def extract_numbers(text: str) -> list:
    if not text:
        return []
    numbers = NUMBER_PATTERN.findall(str(text))
    unique = []
    for num in numbers:
        if num not in unique:
            unique.append(num)
    return unique


def compute_token_metrics(query_tokens: set, candidate_tokens: set):
    if not query_tokens or not candidate_tokens:
        return 0.0, 0.0
    intersection = query_tokens & candidate_tokens
    union = query_tokens | candidate_tokens
    jaccard = len(intersection) / len(union)
    coverage = len(intersection) / len(query_tokens)
    return jaccard, coverage


def compute_number_overlap(query_numbers: set, candidate_numbers: set) -> float:
    if not query_numbers or not candidate_numbers:
        return 0.0
    intersection = query_numbers & candidate_numbers
    return len(intersection) / len(query_numbers)


def compute_similarity_score(query_info: dict, row_info: dict):
    seq_score = SequenceMatcher(None, query_info["normalized"], row_info["normalized"]).ratio()
    token_jaccard, token_coverage = compute_token_metrics(query_info["token_set"], row_info["token_set"])
    number_overlap = compute_number_overlap(query_info["number_set"], row_info["number_set"])
    bonus = 0.0
    if query_info["normalized"] and query_info["normalized"] in row_info["normalized"]:
        bonus += 0.05
    if token_coverage >= 0.75:
        bonus += 0.05
    if number_overlap >= 0.5:
        bonus += 0.05
    if seq_score >= 0.85:
        bonus += 0.05
    base = (seq_score * 0.4) + (token_jaccard * 0.3) + (token_coverage * 0.2) + (number_overlap * 0.1)
    score = min(1.0, base + bonus)
    components = {
        "text": seq_score,
        "token_jaccard": token_jaccard,
        "token_coverage": token_coverage,
        "number": number_overlap,
        "bonus": bonus,
    }
    return score, components


def load_database_rows():
    if not os.path.exists(DB_PATH):
        return []
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cols = {
            row[1].strip().lower()
            for row in conn.execute(f'PRAGMA table_info("{DB_TABLE}")').fetchall()
        }
        has_laying = "cable laying mhr" in cols
        has_term = "cable termination mhr" in cols
        has_mhr = "mhr" in cols
        if has_laying or has_term:
            rows = conn.execute(
                f'SELECT description, "Cable laying Mhr" AS cable_laying, "Cable termination Mhr" AS cable_termination '
                f'FROM "{DB_TABLE}" WHERE description IS NOT NULL'
            ).fetchall()
        elif has_mhr:
            rows = conn.execute(
                f'SELECT description, "Mhr" AS mhr FROM "{DB_TABLE}" WHERE description IS NOT NULL'
            ).fetchall()
        else:
            return []

        data = []
        for row in rows:
            description = row["description"]
            normalized = normalize_description(description)
            tokens = extract_tokens(description)
            numbers = extract_numbers(description)
            cable_laying = row["cable_laying"] if has_laying or has_term else row["mhr"]
            cable_termination = row["cable_termination"] if has_laying or has_term else row["mhr"]
            data.append(
                {
                    "description": description,
                    "cable_laying": cable_laying,
                    "cable_termination": cable_termination,
                    "normalized": normalized,
                    "tokens": tokens,
                    "token_set": set(tokens),
                    "numbers": numbers,
                    "number_set": set(numbers),
                }
            )
        return data
    except sqlite3.Error:
        return []
    finally:
        if conn:
            conn.close()

DATABASE_ROWS = load_database_rows()

def learn_mhr(description: str, value: float) -> bool:
    """
    Learn a new MHR value for a given description.
    Updates existing entry if found, otherwise inserts new one.
    """
    if not description or value is None:
        return False
        
    normalized = normalize_description(description)
    if not normalized:
        return False
        
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Check if description exists
        cursor.execute(f'SELECT count(*) FROM "{DB_TABLE}" WHERE description = ?', (description,))
        count = cursor.fetchone()[0]
        
        cols = {
            row[1].strip().lower()
            for row in conn.execute(f'PRAGMA table_info("{DB_TABLE}")').fetchall()
        }
        has_laying = "cable laying mhr" in cols
        has_term = "cable termination mhr" in cols
        has_mhr = "mhr" in cols
        
        if count > 0:
            # Update existing
            if has_laying:
                cursor.execute(f'UPDATE "{DB_TABLE}" SET "Cable laying Mhr" = ? WHERE description = ?', (value, description))
            elif has_mhr:
                cursor.execute(f'UPDATE "{DB_TABLE}" SET "Mhr" = ? WHERE description = ?', (value, description))
        else:
            # Insert new
            if has_laying:
                cursor.execute(
                    f'INSERT INTO "{DB_TABLE}" (description, "Cable laying Mhr", "Cable termination Mhr") VALUES (?, ?, ?)',
                    (description, value, 0)
                )
            elif has_mhr:
                cursor.execute(
                    f'INSERT INTO "{DB_TABLE}" (description, "Mhr") VALUES (?, ?)',
                    (description, value)
                )
        
        conn.commit()
        
        # Reload cache
        global DATABASE_ROWS
        DATABASE_ROWS = load_database_rows()
        return True
        
    except Exception as e:
        print(f"Error learning MHR: {e}")
        return False
    finally:
        if conn:
            conn.close()


def format_mhr_for_prompt(value):
    if value is None:
        return "N/A"
    if isinstance(value, (int, float)):
        return f"{value:.3f}"
    return str(value)

def get_candidate_rows(description: str, limit: int = MAX_DB_CANDIDATES):
    query_str = "" if description is None else str(description)
    query_info = {
        "original": query_str,
        "normalized": normalize_description(query_str),
    }
    query_tokens = extract_tokens(query_str)
    if not query_tokens and query_info["normalized"]:
        query_tokens = query_info["normalized"].split()
    query_numbers = extract_numbers(query_str)
    query_info["tokens"] = query_tokens
    query_info["token_set"] = set(query_tokens)
    query_info["numbers"] = query_numbers
    query_info["number_set"] = set(query_numbers)

    if not query_str.strip() or not DATABASE_ROWS:
        return [], query_info

    scored = []
    for row in DATABASE_ROWS:
        score, components = compute_similarity_score(query_info, row)
        shared_tokens = [token for token in row["tokens"] if token in query_info["token_set"]][:20]
        shared_numbers = [num for num in row["numbers"] if num in query_info["number_set"]][:20]
        scored.append((score, row, components, shared_tokens, shared_numbers))

    scored.sort(key=lambda item: item[0], reverse=True)

    candidates = []
    for score, row, components, shared_tokens, shared_numbers in scored:
        if score < MIN_CANDIDATE_SCORE and len(candidates) >= MIN_RESULTS_BEFORE_THRESHOLD:
            break
        candidate = {
            "description": row["description"],
            "cable_laying": row["cable_laying"],
            "cable_termination": row["cable_termination"],
            "normalized": row["normalized"],
            "tokens": row["tokens"],
            "numbers": row["numbers"],
            "shared_tokens": shared_tokens,
            "shared_numbers": shared_numbers,
            "score": round(score, 3),
            "score_details": (
                f"text={components['text']:.3f}, token_jaccard={components['token_jaccard']:.3f}, "
                f"coverage={components['token_coverage']:.3f}, numbers={components['number']:.3f}, bonus={components['bonus']:.3f}"
            ),
        }
        candidates.append(candidate)
        if len(candidates) >= limit:
            break

    if not candidates and scored:
        score, row, components, shared_tokens, shared_numbers = scored[0]
        candidates.append({
            "description": row["description"],
            "cable_laying": row["cable_laying"],
            "cable_termination": row["cable_termination"],
            "normalized": row["normalized"],
            "tokens": row["tokens"],
            "numbers": row["numbers"],
            "shared_tokens": shared_tokens,
            "shared_numbers": shared_numbers,
            "score": round(score, 3),
            "score_details": (
                f"text={components['text']:.3f}, token_jaccard={components['token_jaccard']:.3f}, "
                f"coverage={components['token_coverage']:.3f}, numbers={components['number']:.3f}, bonus={components['bonus']:.3f}"
            ),
        })

    return candidates[:limit], query_info


def is_numeric_string(s: str) -> bool:
    if s is None:
        return False
    s = s.strip()
    return True if re.fullmatch(r"\d+(\.\d+)?", s) else False


def to_numeric_string(value) -> str:
    if isinstance(value, (int, float)):
        text = f"{value:.3f}".rstrip("0").rstrip(".")
        return text if text else "0"
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if is_numeric_string(stripped) else ""
    return ""


def fallback_local_mhr(candidates):
    for row in candidates:
        for key in ("cable_laying", "cable_termination"):
            numeric = to_numeric_string(row.get(key))
            if numeric:
                return numeric
    return ""


def find_header_row_and_cols(ws, search_rows=30):
    """
    Find the header row containing a cell == 'description' (case-insensitive).
    Return (header_row_index, desc_col_index). Both are 1-based indices for openpyxl.
    If not found, returns (None, None).
    """
    for r in range(1, min(ws.max_row, search_rows) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip().lower() == "description":
                return r, c
    return None, None

def find_or_create_mhr_column(ws, header_row, desc_col):
    """
    Find Mhr column in the header row. If exists, return it.
    Else if the cell immediately right of Description is empty, use it for Mhr.
    Else scan to the right for the first empty cell in the header row and set it to 'Mhr'.
    Return the selected column index (1-based). Never inserts columns (no shifting).
    """
    # 1) Existing "Mhr" header anywhere in the header row?
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if isinstance(val, str) and val.strip().lower() == "mhr":
            return c

    # 2) If cell immediately to the right of Description is empty, use it
    right_c = desc_col + 1
    # Expand max_column if we will write beyond current max (openpyxl can handle)
    if ws.cell(row=header_row, column=right_c).value in (None, ""):
        ws.cell(row=header_row, column=right_c, value="Mhr")
        return right_c

    # 3) Otherwise find the first empty cell to the right of Description in that header row
    c = desc_col + 1
    while True:
        val = ws.cell(row=header_row, column=c).value
        if val in (None, ""):
            ws.cell(row=header_row, column=c, value="Mhr")
            return c
        c += 1  # keep moving right without inserting columns




def get_model_mhr(description: str) -> str:
    """
    Use GPT-4.1 mini with local database context to select an Mhr value.
    """
    candidates, query_info = get_candidate_rows(description)
    if not candidates:
        return ""

    prompt_candidates = candidates[:MAX_GPT_CONTEXT]
    candidate_lines = []
    query_tokens = query_info.get("tokens") or []
    query_numbers = query_info.get("numbers") or []
    for idx, row in enumerate(prompt_candidates, 1):
        candidate_tokens = row.get("tokens", [])
        candidate_numbers = row.get("numbers", [])
        shared_tokens_list = row.get("shared_tokens", [])
        shared_numbers_list = row.get("shared_numbers", [])
        tokens_text = ", ".join(candidate_tokens[:12]) or "-"
        shared_tokens_text = ", ".join(shared_tokens_list[:12]) or "-"
        missing_tokens_text = ", ".join([tok for tok in query_tokens if tok not in shared_tokens_list][:12]) or "-"
        numbers_text = ", ".join(candidate_numbers[:12]) or "-"
        shared_numbers_text = ", ".join(shared_numbers_list[:12]) or "-"
        missing_numbers_text = ", ".join([num for num in query_numbers if num not in shared_numbers_list][:12]) or "-"
        candidate_lines.append(
            "\n".join(
                [
                    f"{idx}. Description: {row.get('description')}",
                    f"   Normalized candidate: {row.get('normalized')}",
                    f"   Similarity score: {row.get('score')} ({row.get('score_details')})",
                    f"   Cable laying Mhr: {format_mhr_for_prompt(row.get('cable_laying'))}",
                    f"   Cable termination Mhr: {format_mhr_for_prompt(row.get('cable_termination'))}",
                    f"   Shared tokens: {shared_tokens_text}",
                    f"   Missing query tokens: {missing_tokens_text}",
                    f"   Candidate tokens: {tokens_text}",
                    f"   Shared numbers: {shared_numbers_text}",
                    f"   Missing query numbers: {missing_numbers_text}",
                    f"   Candidate numbers: {numbers_text}",
                ]
            )
        )
    candidate_text = "\n".join(candidate_lines)

    query_tokens_text = ", ".join(query_tokens[:12]) or "-"
    query_numbers_text = ", ".join(query_numbers[:12]) or "-"
    query_summary_lines = [
        f"Input description: {query_info.get('original')}",
        f"Normalized: {query_info.get('normalized')}",
        f"Tokens: {query_tokens_text}",
        f"Numbers: {query_numbers_text}",
    ]
    query_summary = "\n".join(query_summary_lines)

    try:
        resp = client.chat.completions.create(
            model=GPT_MODEL,
            temperature=0,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are Electrical Estimator. You receive a BOM description and candidate rows from a local database with similarity details, tokens, and numbers. "
                        "Choose the most appropriate man-hour (Mhr) value. Prefer the Cable laying Mhr when termination isnt in description else use the Cable termination Mhr. "
                        "Return ONLY the numeric value (e.g., 0.450). If none of the candidates apply, return an empty string."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        f"{query_summary}\n\nCandidate rows (best match first):\n{candidate_text}"
                    ),
                },
            ],
        )
        out = (resp.choices[0].message.content or "").strip()
        if is_numeric_string(out):
            return out
    except Exception:
        pass

    return fallback_local_mhr(candidates)


def process_one_sheet(ws, status_callback=None, progress_callback=None, total_rows=0, processed_count=0):
    """
    Process a single worksheet IN-PLACE without changing layout.
    1) Find header row and 'description' column.
    2) Choose/create 'Mhr' column without inserting columns.
    3) For each data row, call the model and write numeric Mhr or blank.
    """
    header_row, desc_col = find_header_row_and_cols(ws)
    if not header_row or not desc_col:
        # No description header → skip this sheet
        if status_callback:
            status_callback(f"Skip: no 'description' header found")
        return 0

    mhr_col = find_or_create_mhr_column(ws, header_row, desc_col)
    start_row = header_row + 1
    end_row = ws.max_row
    row_count = end_row - start_row + 1

    if status_callback:
        status_callback(
            f"Header row {header_row}, Desc col {get_column_letter(desc_col)}, "
            f"Mhr col {get_column_letter(mhr_col)}"
        )

    # Loop through rows and write Mhr
    for r in range(start_row, end_row + 1):
        desc_val = ws.cell(row=r, column=desc_col).value
        if desc_val is None or str(desc_val).strip() == "":
            # leave Mhr blank if there's no description
            ws.cell(row=r, column=mhr_col, value="")
        else:
            mhr = get_model_mhr(str(desc_val))
            ws.cell(row=r, column=mhr_col, value=mhr)
        
        # Update progress
        if progress_callback:
            progress_callback(processed_count + (r - start_row + 1), total_rows)
    
    return row_count

class ModernButton(tk.Frame):
    """Custom modern button with hover effects - FIXED VERSION"""
    def __init__(self, parent, text, command, color=COLORS["primary"], width=120, height=40):
        super().__init__(parent, bg=COLORS["dark_bg"])
        self.command = command
        self.color = color
        self.hover_color = self._adjust_color(color, 20)  # Lighten on hover
        self.width = width
        self.height = height
        
        # Create a label that will act as our button
        self.btn_label = tk.Label(
            self, 
            text=text, 
            bg=color,
            fg=COLORS["text_primary"],
            font=("Segoe UI", 10, "bold"),
            width=width//8,  # Approximate width based on character count
            height=height//20,
            relief="flat",
            cursor="hand2"
        )
        self.btn_label.pack(padx=2, pady=2)
        
        # Bind events
        self.btn_label.bind("<Button-1>", self._on_click)
        self.btn_label.bind("<Enter>", self._on_enter)
        self.btn_label.bind("<Leave>", self._on_leave)
        
    def _on_click(self, event):
        self.command()
        
    def _on_enter(self, event):
        self.btn_label.configure(bg=self.hover_color)
        
    def _on_leave(self, event):
        self.btn_label.configure(bg=self.color)
        
    def _adjust_color(self, hex_color, amount):
        """Lighten or darken a color by a given amount"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        
        new_rgb = []
        for channel in rgb:
            new_channel = max(0, min(255, channel + amount))
            new_rgb.append(new_channel)
            
        return f"#{new_rgb[0]:02x}{new_rgb[1]:02x}{new_rgb[2]:02x}"

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Electrical Estimator – Modern Edition")
        self.geometry("800x600")
        self.configure(bg=COLORS["dark_bg"])
        self.resizable(True, True)
        
        # Set application icon (if available)
        try:
            self.iconbitmap("icon.ico")  # Optional: add an icon file
        except:
            pass

        self.filepath = None
        self.sheets = []
        self.total_rows = 0
        self.processed_rows = 0
        
        # Configure styles
        self._configure_styles()
        self._build_ui()

    def _configure_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        
        # Configure styles
        style.configure("TFrame", background=COLORS["dark_bg"])
        style.configure("TLabel", background=COLORS["dark_bg"], foreground=COLORS["text_primary"])
        style.configure("Card.TFrame", background=COLORS["card_bg"], relief="raised", borderwidth=1)
        style.configure("Title.TLabel", background=COLORS["card_bg"], foreground=COLORS["text_primary"], font=("Segoe UI", 12, "bold"))
        style.configure("Subtitle.TLabel", background=COLORS["dark_bg"], foreground=COLORS["text_secondary"], font=("Segoe UI", 10))
        style.configure("TButton", padding=8, background=COLORS["primary"], foreground=COLORS["text_primary"])
        style.map("TButton", background=[("active", COLORS["primary"])])
        style.configure("TCheckbutton", background=COLORS["dark_bg"], foreground=COLORS["text_primary"])
        style.configure("TCombobox", fieldbackground=COLORS["light_bg"], background=COLORS["light_bg"], 
                        foreground=COLORS["text_primary"], selectbackground=COLORS["primary"])
        style.configure("Horizontal.TProgressbar", troughcolor=COLORS["light_bg"], background=COLORS["success"])

    def _build_ui(self):
        # Main container with padding
        main_container = ttk.Frame(self, padding=20)
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        title = ttk.Label(
            header_frame, 
            text="Electrical Estimator", 
            font=("Segoe UI", 18, "bold"),
            foreground=COLORS["text_primary"]
        )
        title.pack(side=tk.LEFT)
        
        subtitle = ttk.Label(
            header_frame, 
            text="AI-Powered Mhr Estimation", 
            font=("Segoe UI", 12),
            foreground=COLORS["text_secondary"]
        )
        subtitle.pack(side=tk.LEFT, padx=(10, 0), pady=(4, 0))
        
        # ================= UNIFIED SINGLE CARD FOR STEPS 1–3 =================
        workflow_card = ttk.Frame(main_container, style="Card.TFrame", padding=15)
        workflow_card.pack(fill=tk.X, pady=(0, 15))

        # Step 1: File selection
        ttk.Label(workflow_card, text="1. Select BOM File", style="Title.TLabel").pack(anchor="w", pady=(0, 10))
        
        file_selector_frame = ttk.Frame(workflow_card)
        file_selector_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.lbl_file = ttk.Label(
            file_selector_frame, 
            text="No file selected", 
            foreground=COLORS["text_secondary"],
            font=("Segoe UI", 9)
        )
        self.lbl_file.pack(side=tk.LEFT, fill=tk.X, expand=True, anchor="w")
        
        browse_btn = ModernButton(
            file_selector_frame, 
            "Browse Files", 
            self._browse_file,
            color=COLORS["primary"],
            width=120
        )
        browse_btn.pack(side=tk.RIGHT)
        
        # Step 2: Sheet selection
        ttk.Label(workflow_card, text="2. Sheet Selection", style="Title.TLabel").pack(anchor="w", pady=(15, 10))
        
        sheet_selector_frame = ttk.Frame(workflow_card)
        sheet_selector_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(
            sheet_selector_frame, 
            text="Select worksheet:", 
            foreground=COLORS["text_primary"]
        ).pack(side=tk.LEFT, anchor="w")
        
        self.cmb_sheet = ttk.Combobox(
            sheet_selector_frame, 
            state="readonly", 
            values=self.sheets, 
            width=30,
            font=("Segoe UI", 10)
        )
        self.cmb_sheet.pack(side=tk.LEFT, padx=(10, 0))
        
        self.var_all = tk.BooleanVar(value=False)
        all_sheets_check = ttk.Checkbutton(
            sheet_selector_frame, 
            text="Process ALL sheets", 
            variable=self.var_all, 
            command=self._toggle_all,
            style="TCheckbutton"
        )
        all_sheets_check.pack(side=tk.RIGHT)
        
        # Step 3: Process
        ttk.Label(workflow_card, text="3. Process & Save", style="Title.TLabel").pack(anchor="w", pady=(15, 10))
        
        process_btn = ModernButton(
            workflow_card, 
            "Process & Save Results", 
            self._start_process,
            color=COLORS["success"],
            width=200
        )
        process_btn.pack(pady=5)
        # =====================================================================
        
        # Progress section
        progress_frame = ttk.Frame(main_container)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        progress_header = ttk.Frame(progress_frame)
        progress_header.pack(fill=tk.X)
        
        ttk.Label(
            progress_header, 
            text="Progress", 
            font=("Segoe UI", 11, "bold"),
            foreground=COLORS["text_primary"]
        ).pack(side=tk.LEFT, anchor="w", pady=(0, 5))
        
        self.progress_text = ttk.Label(
            progress_header,
            text="0%",
            foreground=COLORS["text_secondary"],
            font=("Segoe UI", 10)
        )
        self.progress_text.pack(side=tk.RIGHT, anchor="e", pady=(0, 5))
        
        self.pbar = ttk.Progressbar(
            progress_frame, 
            orient="horizontal", 
            mode="determinate", 
            maximum=100,
            style="Horizontal.TProgressbar"
        )
        self.pbar.pack(fill=tk.X, pady=(0, 10))
        
        # Status area
        status_frame = ttk.Frame(main_container)
        status_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(
            status_frame, 
            text="Status", 
            font=("Segoe UI", 11, "bold"),
            foreground=COLORS["text_primary"]
        ).pack(anchor="w", pady=(0, 5))
        
        # Status text area with scrollbar
        status_container = ttk.Frame(status_frame)
        status_container.pack(fill=tk.BOTH, expand=True)
        
        self.status_text = tk.Text(
            status_container,
            height=8,
            bg=COLORS["light_bg"],
            fg=COLORS["text_primary"],
            font=("Consolas", 9),
            wrap=tk.WORD,
            relief="flat",
            padx=10,
            pady=10,
            borderwidth=1,
            highlightthickness=1,
            highlightcolor=COLORS["border"],
            highlightbackground=COLORS["border"]
        )
        
        scrollbar = ttk.Scrollbar(status_container, orient="vertical", command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=scrollbar.set)
        
        self.status_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Make text widget read-only
        self.status_text.configure(state="disabled")
        
        # Footer
        footer_frame = ttk.Frame(main_container)
        footer_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(
            footer_frame, 
            text="AI Electrical Estimator v2.0", 
            foreground=COLORS["text_secondary"],
            font=("Segoe UI", 9)
        ).pack(side=tk.LEFT)
        
        ttk.Label(
            footer_frame, 
            text="Powered by OpenAI", 
            foreground=COLORS["text_secondary"],
            font=("Segoe UI", 9)
        ).pack(side=tk.RIGHT)

    def _toggle_all(self):
        self.cmb_sheet.configure(state=("disabled" if self.var_all.get() else "readonly"))

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel file", 
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return
        self.filepath = path
        self.lbl_file.configure(text=os.path.basename(path))
        try:
            wb = load_workbook(path)
            self.sheets = wb.sheetnames
            self.cmb_sheet.configure(values=self.sheets)
            if self.sheets:
                self.cmb_sheet.set(self.sheets[0])
                
            # Update status
            self._update_status(f"Loaded: {os.path.basename(path)}")
            self._update_status(f"Found {len(self.sheets)} sheet(s): {', '.join(self.sheets)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open workbook:\n{e}")

    def _start_process(self):
        if not self.filepath:
            messagebox.showwarning("Missing file", "Please choose a BOM .xlsx file first.")
            return
        if not self.var_all.get() and not self.cmb_sheet.get():
            messagebox.showwarning("Missing sheet", "Pick a sheet or check 'Process ALL sheets'.")
            return

        default_name = f"processed_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Save processed Excel", 
            defaultextension=".xlsx", 
            initialfile=default_name, 
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return

        # Reset progress bar
        self.pbar["value"] = 0
        self.progress_text["text"] = "0%"
        
        t = threading.Thread(target=self._run_process, args=(save_path,), daemon=True)
        t.start()

    def _update_progress(self, current, total):
        """Update progress bar with current progress"""
        if total > 0:
            progress_percent = (current / total) * 100
            self.pbar["value"] = progress_percent
            self.progress_text["text"] = f"{progress_percent:.1f}%"
        self.update_idletasks()

    def _run_process(self, save_path: str):
        try:
            wb = load_workbook(self.filepath)
            target_sheets = wb.sheetnames if self.var_all.get() else [self.cmb_sheet.get()]
            
            # Calculate total rows to process for progress tracking
            self.total_rows = 0
            for name in target_sheets:
                ws = wb[name]
                header_row, desc_col = find_header_row_and_cols(ws)
                if header_row and desc_col:
                    self.total_rows += ws.max_row - header_row
            
            self._update_status(f"Starting processing of {len(target_sheets)} sheet(s), {self.total_rows} rows...")
            
            processed_rows = 0
            for name in target_sheets:
                self._update_status(f"Processing sheet: {name}")
                ws = wb[name]
                rows_processed = process_one_sheet(
                    ws, 
                    status_callback=self._update_status, 
                    progress_callback=self._update_progress,
                    total_rows=self.total_rows,
                    processed_count=processed_rows
                )
                processed_rows += rows_processed
                self._update_status(f"Completed: {name} ({rows_processed} rows)")

            wb.save(save_path)
            self._update_progress(self.total_rows, self.total_rows)  # Set to 100%
            self._update_status(f"Successfully processed {processed_rows} rows across {len(target_sheets)} sheet(s)")
            self._update_status(f"Saved to: {save_path}")
            messagebox.showinfo("Success", f"Processed file saved to:\n{save_path}")
        except Exception as e:
            self.pbar["value"] = 0
            self.progress_text["text"] = "0%"
            self._update_status(f"Error: {str(e)}")
            messagebox.showerror("Processing failed", str(e))

    def _update_status(self, text: str):
        """Update status text area with new message"""
        self.status_text.configure(state="normal")
        self.status_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {text}\n")
        self.status_text.see(tk.END)  # Auto-scroll to bottom
        self.status_text.configure(state="disabled")
        self.update_idletasks()


if __name__ == "__main__":
    # Quick sanity check for API key/model so the user sees a clear error if they forgot to set them
    if not OPENAI_API_KEY or OPENAI_API_KEY == "YOUR_OPENAI_API_KEY":
        messagebox.showerror("Missing API key", "Please open the script and set OPENAI_API_KEY.")
    elif not os.path.exists(DB_PATH):
        messagebox.showerror("Missing database", f"Expected local SQLite database at: {DB_PATH}")
    else:
        app = App()
        app.mainloop()
